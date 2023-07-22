import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import shutil

#loading the excel file 
Workbook = openpyxl.load_workbook(r"C:\Users\Dell\Desktop\Task\IN\formate.xlsx") 
worksheet = Workbook['Sheet1']

row = worksheet.max_row
column = worksheet.max_column

#looping through rows and column
for i in range(1, row + 1):
    for j in range(1, column + 1):
        if worksheet.cell(i, j).value < 0:
            worksheet.cell(i, j).fill = PatternFill(patternType="solid", fgColor="f20a0a")
        elif worksheet.cell(i, j).value > 0:
           worksheet.cell(i, j).fill  = PatternFill(patternType="solid", fgColor="2df20a")
        else:
            worksheet.cell(i, j).fill = PatternFill(patternType="solid", fgColor="f2ee0a")
#saving the workbook
Workbook.save("formatted_data.xlsx")

#moving file from one directory to another 
source_path = r"C:\Users\Dell\Desktop\Task\IN\formate.xlsx"
dest_path = r"C:\Users\Dell\Desktop\Task\Archive\formate.xlsx"
try:
    shutil.move(source_path, dest_path)
    print('file moved succesfully')
except IOError:
    print('file already exsist')
